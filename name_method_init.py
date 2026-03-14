#!/usr/bin/env python3
"""
命名方法配置工具
允许用户交互式选择 Excel 表格的命名方式：
- 行命名：每行对应一个人，使用指定列的内容作为前缀
- 列命名：每列对应一个人，使用指定行的内容作为前缀
"""

import json
import os
from openpyxl import load_workbook


class NamingMethodConfig:
    """命名方法配置类"""
    
    def __init__(self):
        self.config_file = "naming_config.json"
        self.config = {}
    
    def load_config(self):
        """加载现有的配置"""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    self.config = json.load(f)
                return True
            except Exception as e:
                print(f"加载配置失败: {e}")
                return False
        return False
    
    def save_config(self):
        """保存配置到文件"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
            print(f"配置已保存到 {self.config_file}")
            return True
        except Exception as e:
            print(f"保存配置失败: {e}")
            return False
    
    def preview_excel_structure(self, excel_path):
        """预览 Excel 表格结构，帮助用户选择命名方式"""
        try:
            wb = load_workbook(excel_path, data_only=True)
            sheet = wb.active
            
            print(f"\n=== Excel 表格预览 ({excel_path}) ===")
            print(f"工作表: {sheet.title}")
            print(f"行数: {sheet.max_row}")
            print(f"列数: {sheet.max_column}")
            
            # 显示前几行和列的示例数据
            print("\n前5行前5列的数据预览:")
            for row in range(1, min(6, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(6, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value)[:20] if cell_value is not None else "")
                print(f"行{row}: {' | '.join(row_data)}")
            
            return True
        except Exception as e:
            print(f"预览 Excel 失败: {e}")
            return False
    
    def choose_naming_method(self):
        """选择命名方法"""
        print("\n=== 选择命名方法 ===")
        print("1. 行命名 - 每行对应一个人，使用指定列的内容作为文件名")
        print("2. 列命名 - 每列对应一个人，使用指定行的内容作为文件名")
        
        while True:
            choice = input("请选择命名方法 (1/2): ").strip()
            if choice == "1":
                return "row"
            elif choice == "2":
                return "col"
            else:
                print("无效选择，请输入 1 或 2")
    
    def configure_row_naming(self, excel_path):
        """配置行命名方法"""
        print("\n=== 配置行命名方法 ===")
        print("说明：每行对应一个人，使用指定列的内容作为文件名前缀")
        
        try:
            wb = load_workbook(excel_path, data_only=True)
            sheet = wb.active
            
            # 显示列标题（第一行）
            print("\n列标题（第一行）:")
            for col in range(1, min(11, sheet.max_column + 1)):
                cell_value = sheet.cell(row=1, column=col).value
                print(f"列{col}: {cell_value}")
            
            # 选择姓名列
            while True:
                try:
                    name_col = int(input("\n请输入姓名所在的列号 (例如：列1输入1, 列2输入2...): ").strip())
                    if 1 <= name_col <= sheet.max_column:
                        break
                    else:
                        print(f"列号必须在 1-{sheet.max_column} 之间")
                except ValueError:
                    print("请输入有效的数字")
            
            # 验证姓名列
            print(f"\n验证姓名列（第{name_col}列）的前几行数据:")
            for row in range(1, min(6, sheet.max_row + 1)):
                name = sheet.cell(row=row, column=name_col).value
                print(f"行{row}: {name}")
            
            confirm = input("\n确认使用此列作为姓名列？(y/n): ").strip().lower()
            if confirm != 'y':
                print("配置已取消")
                return None
            
            # 配置完成
            config = {
                "method": "row",
                "name_column": name_col,
                "separator": "-"
            }
            
            print("\n行命名配置完成:")
            print(f"- 命名方法: 行命名")
            print(f"- 姓名列: 第{name_col}列")
            print(f"- 文件名格式: 姓名-序号 (例如: 张三-1, 张三-2)")
            
            return config
            
        except Exception as e:
            print(f"配置行命名失败: {e}")
            return None
    
    def configure_column_naming(self, excel_path):
        """配置列命名方法"""
        print("\n=== 配置列命名方法 ===")
        print("说明：每列对应一个人，使用指定行的内容作为文件名前缀")
        
        try:
            wb = load_workbook(excel_path, data_only=True)
            sheet = wb.active
            
            # 显示行数据（第一列）
            print("\n行数据（第一列）:")
            for row in range(1, min(11, sheet.max_row + 1)):
                cell_value = sheet.cell(row=row, column=1).value
                print(f"行{row}: {cell_value}")
            
            # 选择姓名行
            while True:
                try:
                    name_row = int(input("\n请输入姓名所在的行号 (例如：第一行输入1, 第二行输入2): ").strip())
                    if 1 <= name_row <= sheet.max_row:
                        break
                    else:
                        print(f"行号必须在 1-{sheet.max_row} 之间")
                except ValueError:
                    print("请输入有效的数字")
            
            # 验证姓名行
            print(f"\n验证姓名行（第{name_row}行）的前几列数据:")
            for col in range(1, min(11, sheet.max_column + 1)):
                name = sheet.cell(row=name_row, column=col).value
                print(f"列{col}: {name}")
            
            confirm = input("\n确认使用此行作为姓名行？(y/n): ").strip().lower()
            if confirm != 'y':
                print("配置已取消")
                return None
            
            # 配置完成
            config = {
                "method": "col",
                "name_row": name_row,
                "separator": "-"
            }
            
            print("\n列命名配置完成:")
            print(f"- 命名方法: 列命名")
            print(f"- 姓名行: 第{name_row}行")
            print(f"- 文件名格式: 姓名-序号 (例如: 张三-1, 李四-1)")
            
            return config
            
        except Exception as e:
            print(f"配置列命名失败: {e}")
            return None
    
    def run_interactive_config(self):
        """运行交互式配置"""
        print("=== Excel 文件命名方法配置工具 ===")
        
        # 获取 Excel 文件路径
        excel_path = ""
        
        if not excel_path:
            import glob
            xlsx_files = glob.glob("*.xlsx")
            if not xlsx_files:
                print("当前目录下未找到 .xlsx 文件")
                return False
            
            # 对文件进行排序：Example.xlsx 放在最后
            def sort_key(filename):
                if filename.lower() == "example.xlsx":
                    return (1, filename)  # Example.xlsx 排最后
                return (0, filename)  # 其他文件排前面
            
            xlsx_files.sort(key=sort_key)
            
            # 显示选择器（即使只有一个文件也显示选择器）
            print("\n检测到 .xlsx 文件，请选择一个：")
            print("0- 手动输入/拖入文件路径")
            for i, f in enumerate(xlsx_files, 1):
                print(f"{i}- {f}")
            
            while True:
                choice = input("\n请输入选项编号: ").strip()
                if choice == "0":
                    excel_path = input("请输入 Excel 文件路径: ").strip()
                    if not excel_path:
                        print("未输入路径，退出。")
                        return False
                    # 去除拖入路径可能带来的引号
                    excel_path = excel_path.strip('"\'')
                    break
                elif choice.isdigit():
                    idx = int(choice) - 1
                    if 0 <= idx < len(xlsx_files):
                        excel_path = xlsx_files[idx]
                        print(f"已选择: {excel_path}")
                        break
                    else:
                        print(f"无效选项，请输入 0-{len(xlsx_files)}")
                else:
                    print(f"无效选项，请输入 0-{len(xlsx_files)}")
        
        if not os.path.exists(excel_path):
            print(f"文件不存在: {excel_path}")
            return False
        
        # 预览 Excel 结构
        if not self.preview_excel_structure(excel_path):
            return False
        
        # 选择命名方法
        method = self.choose_naming_method()
        
        # 根据选择进行配置
        if method == "row":
            config = self.configure_row_naming(excel_path)
        else:
            config = self.configure_column_naming(excel_path)
        
        if not config:
            return False
        
        # 保存配置
        self.config = config
        self.config["excel_path"] = excel_path
        
        if self.save_config():
            print("\n配置已保存！")
            print("现在可以运行 extract_links.py 来使用新的命名方法提取链接。")
            return True
        else:
            return False


def get_naming_config():
    """获取命名配置（供 extract_links.py 使用）"""
    config_file = "naming_config.json"
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"读取命名配置失败: {e}")
    return None


def generate_filename(config, row, col, url_count, url_index):
    """根据配置生成文件名"""
    if not config:
        return f"unknown_{url_index}"
    
    try:
        wb = load_workbook(config.get("excel_path"), data_only=True)
        sheet = wb.active
        separator = config.get("separator", "-")
        
        if config.get("method") == "row":
            # 行命名：使用指定列的内容作为前缀
            name_col = config.get("name_column", 1)
            name = sheet.cell(row=row, column=name_col).value
            name = str(name).strip() if name is not None else "unknown"
            # 移除文件名中常见非法字符
            import re
            name = re.sub(r'[\\/:"*?<>|]+', "_", name)
            return f"{name}{separator}{url_count}"
        
        elif config.get("method") == "col":
            # 列命名：使用指定行的内容作为前缀
            name_row = config.get("name_row", 1)
            name = sheet.cell(row=name_row, column=col).value
            name = str(name).strip() if name is not None else "unknown"
            # 移除文件名中常见非法字符
            import re
            name = re.sub(r'[\\/:"*?<>|]+', "_", name)
            return f"{name}{separator}{url_count}"
        
    except Exception as e:
        print(f"生成文件名失败: {e}")
    
    return f"unknown_{url_index}"


if __name__ == "__main__":
    config_tool = NamingMethodConfig()
    
    # 检查是否已有配置
    if config_tool.load_config():
        print("检测到现有配置:")
        print(json.dumps(config_tool.config, ensure_ascii=False, indent=2))
        
        reuse = input("\n是否重新配置？(y/n): ").strip().lower()
        if reuse != 'y':
            print("使用现有配置。")
            exit(0)
    
    # 运行交互式配置
    config_tool.run_interactive_config()