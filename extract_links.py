#!/usr/bin/env python3
"""
从 Excel 表格中提取所有下载链接并保存为新的 Excel 和 aria2 文本列表。
可选地验证链接是否可访问（使用 HTTP HEAD 请求）。
"""
import argparse
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse
import requests
from openpyxl import load_workbook
import pandas as pd
from tqdm import tqdm

URL_RE = re.compile(r"https?://[\w\-\.\~:/\?\#\[\]@!\$&'\(\)\*\+,;=%]+")


def find_urls_in_text(text):
    if not text:
        return []
    return URL_RE.findall(str(text))


def is_valid_url_format(url):
    try:
        p = urlparse(url)
        return p.scheme in ("http", "https") and bool(p.netloc)
    except Exception:
        return False


def check_url(url, timeout=8):
    try:
        # Prefer HEAD to avoid downloading body; follow redirects
        r = requests.head(url, allow_redirects=True, timeout=timeout)
        status = r.status_code
        if status >= 400:
            # sometimes servers don't like HEAD, try GET with small timeout
            r = requests.get(url, stream=True, timeout=timeout)
            status = r.status_code
        return True, status
    except Exception as e:
        return False, str(e)


def extract_links_from_workbook(path, naming_config=None):
    wb = load_workbook(path, data_only=True)
    found = []
    
    # 统计每个姓名对应的 URL 数量（用于生成序号）
    name_url_counts = {}
    
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for row in ws.iter_rows():
            for cell in row:
                cell_val = cell.value
                row_idx = cell.row
                col_idx = cell.column
                
                # 根据命名配置生成前缀
                if naming_config:
                    # 使用新的命名系统
                    from name_method_init import generate_filename
                    
                    # 获取姓名标识符（用于计数）
                    if naming_config.get("method") == "row":
                        # 行命名：使用指定列的内容作为标识
                        name_col = naming_config.get("name_column", 1)
                        name = ws.cell(row=row_idx, column=name_col).value
                        name_key = str(name).strip() if name is not None else "unknown"
                    elif naming_config.get("method") == "col":
                        # 列命名：使用指定行的内容作为标识
                        name_row = naming_config.get("name_row", 1)
                        name = ws.cell(row=name_row, column=col_idx).value
                        name_key = str(name).strip() if name is not None else "unknown"
                    else:
                        name_key = "unknown"
                    
                    # 更新计数
                    if name_key not in name_url_counts:
                        name_url_counts[name_key] = 0
                    name_url_counts[name_key] += 1
                    url_count = name_url_counts[name_key]
                    
                    # 生成前缀
                    prefix = generate_filename(naming_config, row_idx, col_idx, url_count, len(found) + 1)
                else:
                    raise RuntimeError("未找到命名配置。请先运行 name_method_init.py 来配置命名方法。")
                
                # 移除文件名中常见非法字符
                prefix = re.sub(r'[\\/:"*?<>|]+', "_", prefix)
                
                # hyperlinks stored in cell.hyperlink (openpyxl)
                if cell.hyperlink and getattr(cell.hyperlink, "target", None):
                    url = cell.hyperlink.target
                    if is_valid_url_format(url):
                        found.append({
                            "sheet": sheetname,
                            "cell": cell.coordinate,
                            "row": row_idx,
                            "col": col_idx,
                            "prefix": prefix,
                            "text": str(cell_val) if cell_val is not None else "",
                            "url": url,
                            "source": "hyperlink",
                        })
                # extract URLs in cell text
                for url in find_urls_in_text(cell_val):
                    if is_valid_url_format(url):
                        found.append({
                            "sheet": sheetname,
                            "cell": cell.coordinate,
                            "row": row_idx,
                            "col": col_idx,
                            "prefix": prefix,
                            "text": str(cell_val) if cell_val is not None else "",
                            "url": url,
                            "source": "text",
                        })
    return found


def dedupe_keep_first(items):
    seen = set()
    out = []
    for it in items:
        u = it["url"]
        if u in seen:
            continue
        seen.add(u)
        out.append(it)
    return out


def main():
    parser = argparse.ArgumentParser(description="从 Excel 提取下载链接并输出为 Excel 与 aria2 文本列表")
    parser.add_argument("input", nargs="?", default=None, help="输入 Excel 文件路径（默认查找当前目录下的第一个 .xlsx 文件）")
    parser.add_argument("-o", "--output", default="download_links.xlsx", help="输出 Excel 文件名")
    parser.add_argument("--aria2", default="aria2_links.txt", help="输出给 aria2 的纯 URL 列表文件名")
    parser.add_argument("--check", action="store_true", help="对链接做快速可达性检查（会发起 HTTP 请求）")
    parser.add_argument("--workers", type=int, default=10, help="并发检查线程数")
    parser.add_argument("--timeout", type=int, default=8, help="HTTP 请求 timeout（秒）")
    args = parser.parse_args()

    import glob
    import os

    inpath = args.input
    if not inpath:
        xlsx_files = glob.glob("*.xlsx")
        if not xlsx_files:
            print("当前目录下未找到 .xlsx 文件，请指定输入文件。")
            return
        
        # 对文件进行排序：Example.xlsx 放在最后
        def sort_key(filename):
            if filename.lower() == "example.xlsx":
                return (1, filename)  # Example.xlsx 排最后
            return (0, filename)  # 其他文件排前面
        
        xlsx_files.sort(key=sort_key)
        
        # 显示选择器（即使只有一个文件也显示选择器）
        print("\n检测到 .xlsx 文件，请选择一个：")
        print("0- 手动输入/拖入文件路径")
        for i, f in enumerate(xlsx_files, 1):
            print(f"{i}- {f}")
        
        while True:
            choice = input("\n请输入选项编号: ").strip()
            if choice == "0":
                inpath = input("请输入 Excel 文件路径: ").strip()
                if not inpath:
                    print("未输入路径，退出。")
                    return
                # 去除拖入路径可能带来的引号
                inpath = inpath.strip('"\'')
                if not os.path.exists(inpath):
                    print(f"文件不存在: {inpath}")
                    return
                break
            elif choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(xlsx_files):
                    inpath = xlsx_files[idx]
                    print(f"已选择: {inpath}")
                    break
                else:
                    print(f"无效选项，请输入 0-{len(xlsx_files)}")
            else:
                print(f"无效选项，请输入 0-{len(xlsx_files)}")

    # 加载命名配置（自动使用）
    from name_method_init import get_naming_config
    naming_config = get_naming_config()
    if not naming_config:
        raise RuntimeError("未找到命名配置文件。请先运行 name_method_init.py 来配置命名方法。")
    print("已加载命名配置文件，使用自定义命名规则。")

    print("正在从 {} 中提取链接...".format(inpath))
    items = extract_links_from_workbook(inpath, naming_config)
    if not items:
        print("未找到任何链接。")
        return
    items = dedupe_keep_first(items)
    print("已提取 {} 条唯一 URL（不含可达性检查）。".format(len(items)))

    # optionally check
    if args.check:
        print("开始对链接进行可达性检查（并发 {}，超时 {}s）...".format(args.workers, args.timeout))
        with ThreadPoolExecutor(max_workers=args.workers) as ex:
            future_to_item = {ex.submit(check_url, it["url"], args.timeout): it for it in items}
            for fut in tqdm(as_completed(future_to_item), total=len(future_to_item)):
                it = future_to_item[fut]
                ok, info = fut.result()
                it["ok"] = ok
                it["status"] = info
    else:
        for it in items:
            it["ok"] = None
            it["status"] = None

    df = pd.DataFrame(items)
    # 输出 Excel
    #df.to_excel(args.output, index=False)
    #print("已保存结果到 {}".format(args.output))

    # 输出 aria2 列表（每个 URL 后面带 out= 文件名）
    def _safe_basename_from_url(url):
        path = urlparse(url).path
        base = os.path.basename(path)
        if not base:
            base = "file"
        # 移除文件名中常见非法字符
        return re.sub(r'[\\/:"*?<>|]+', "_", base)

    def _make_unique_name(name, seen):
        if name not in seen:
            seen[name] = 0
            return name
        seen[name] += 1
        root, ext = os.path.splitext(name)
        return f"{root}_{seen[name]}{ext}"

    sep = "_"
    seen_names = {}
    with open(args.aria2, "w", encoding="utf-8") as f:
        for it in df.to_dict(orient="records"):
            url = it["url"]
            prefix = it.get("prefix", "unknown")
            base = _safe_basename_from_url(url)
            filename = f"{prefix}{sep}{base}"
            filename = _make_unique_name(filename, seen_names)
            f.write(url + "\n")
            f.write(f"  out={filename}\n")
    print("已保存 aria2 列表到 {}（包含 out= 文件名）".format(args.aria2))


if __name__ == "__main__":
    main()
